"""Excel logging for IABS detection events."""

from __future__ import annotations

import threading
from pathlib import Path
from typing import Any, Protocol

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from settings import EXCEL_FILE_NAME, log_exception_details, write_log


HEADERS = (
    "Tarih",
    "Saat",
    "Kisi Sayisi",
    "Fotograf Adi",
    "Bildirim Turu",
    "Dosya Yolu",
)
HEADER_FILL = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"
DEFAULT_COLUMN_WIDTH = 22
PHOTO_PATH_COLUMN_WIDTH = 58

_EXCEL_LOCK = threading.Lock()


class DetectionEventLike(Protocol):
    """Protocol for detection event values used by Excel logging."""

    date_text: str
    time_text: str
    person_count: int
    photo_name: str
    photo_path: str
    notification_type: str


def write_detection_row(
    event: DetectionEventLike,
    config: dict[str, Any],
) -> bool:
    """Append a detection event to the configured Excel workbook."""
    log_folder = config["folders"]["logs"]

    try:
        with _EXCEL_LOCK:
            workbook_path = _get_workbook_path(config)
            workbook_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = _load_or_create_workbook(workbook_path)
            worksheet = workbook.active
            worksheet.append(
                [
                    event.date_text,
                    event.time_text,
                    event.person_count,
                    event.photo_name,
                    event.notification_type,
                    event.photo_path,
                ]
            )
            _format_worksheet(worksheet)
            workbook.save(workbook_path)

        write_log("Excel kaydi eklendi", log_folder)
        return True
    except Exception as exc:
        log_exception_details(
            exc,
            context="Excel kaydi eklenemedi",
            log_folder=log_folder,
        )
        return False


def _get_workbook_path(config: dict[str, Any]) -> Path:
    """Return the configured Excel workbook path."""
    file_name = str(config["runtime"].get("excel_file_name", EXCEL_FILE_NAME))
    return Path(config["folders"]["excel"]) / file_name


def _load_or_create_workbook(workbook_path: Path) -> Workbook:
    """Open an existing workbook or create a new one with headers."""
    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
        worksheet = workbook.active
        if worksheet.max_row == 0:
            worksheet.append(list(HEADERS))
        return workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "IABS Kayitlari"
    worksheet.append(list(HEADERS))
    _format_worksheet(worksheet)
    workbook.save(workbook_path)
    return workbook


def _format_worksheet(worksheet: Any) -> None:
    """Apply readable formatting to the worksheet."""
    header_fill = PatternFill(
        fill_type="solid",
        start_color=HEADER_FILL,
        end_color=HEADER_FILL,
    )
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column_index in range(1, len(HEADERS) + 1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = DEFAULT_COLUMN_WIDTH

    worksheet.column_dimensions["F"].width = PHOTO_PATH_COLUMN_WIDTH
    worksheet.freeze_panes = "A2"
